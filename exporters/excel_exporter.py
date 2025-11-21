"""
Excel/CSV export functionality
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
from typing import List, Dict
import csv
import config
from database import db


class ExcelExporter:
    """Export generated content to Excel format"""
    
    def __init__(self):
        self.headers = [
            '章节ID',
            '章节名',
            '原文片段',
            '题目类型',
            '题干',
            '选项',
            '答案',
            '解析',
            '生成模型',
            '校验状态'
        ]
    
    def export_book(self, book_id: int, output_path: Path = None) -> Path:
        """
        Export all verified content for a book to Excel
        
        Args:
            book_id: Book ID
            output_path: Output file path (optional)
        
        Returns:
            Path to exported file
        """
        # Get book info
        book = db.get_book_by_id(book_id)
        if not book:
            raise ValueError(f"Book {book_id} not found")
        
        # Generate default output path if not provided
        if not output_path:
            output_path = config.EXPORT_DIR / f"{book['title']}_export.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Generated Content"
        
        # Write headers
        self._write_headers(ws)
        
        # Get all chapters
        chapters = db.get_chapters_by_book(book_id)
        
        row = 2  # Start from row 2 (after headers)
        
        for chapter in chapters:
            # Get generated content for this chapter
            content_list = db.get_generated_content_by_chapter(chapter['id'])
            
            for content in content_list:
                ws.cell(row, 1, chapter['id'])
                ws.cell(row, 2, chapter['title'])
                
                # Content excerpt (first 100 chars)
                content_excerpt = chapter.get('content_md', '')[:100] + '...' if chapter.get('content_md') else ''
                ws.cell(row, 3, content_excerpt)
                
                # Content type
                content_type_cn = '问答' if content['content_type'] == 'qa' else '习题'
                ws.cell(row, 4, content_type_cn)
                
                # Question/题干
                ws.cell(row, 5, content['question'])
                
                # Options (for choice questions)
                import json
                options_str = ''
                if content.get('options_json'):
                    try:
                        options = json.loads(content['options_json'])
                        options_str = '\n'.join(options)
                    except:
                        options_str = content['options_json']
                ws.cell(row, 6, options_str)
                
                # Answer
                ws.cell(row, 7, content['answer'])
                
                # Explanation
                ws.cell(row, 8, content.get('explanation', ''))
                
                # Model
                model_info = f"{content['model_name']}"
                if content.get('model_version'):
                    model_info += f" ({content['model_version']})"
                ws.cell(row, 9, model_info)
                
                # Status
                status_cn = {
                    'pending': '待生成',
                    'generated': '已生成',
                    'verified': '已校验'
                }.get(content['status'], content['status'])
                ws.cell(row, 10, status_cn)
                
                row += 1
        
        # Adjust column widths
        self._adjust_column_widths(ws)
        
        # Save workbook
        wb.save(output_path)
        
        return output_path
    
    def _write_headers(self, ws):
        """Write and style header row"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(self.headers, start=1):
            cell = ws.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _adjust_column_widths(self, ws):
        """Adjust column widths based on content"""
        column_widths = {
            'A': 10,   # 章节ID
            'B': 20,   # 章节名
            'C': 30,   # 原文片段
            'D': 12,    # 题目类型
            'E': 50,   # 题干
            'F': 30,   # 选项
            'G': 30,   # 答案
            'H': 50,   # 解析
            'I': 20,   # 生成模型
            'J': 12    # 校验状态
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def export_to_csv(self, book_id: int, output_path: Path = None) -> Path:
        """
        Export to CSV format
        
        Args:
            book_id: Book ID
            output_path: Output file path (optional)
        
        Returns:
            Path to exported file
        """
        # Get book info
        book = db.get_book_by_id(book_id)
        if not book:
            raise ValueError(f"Book {book_id} not found")
        
        # Generate default output path if not provided
        if not output_path:
            output_path = config.EXPORT_DIR / f"{book['title']}_export.csv"
        
        # Get all chapters and content
        chapters = db.get_chapters_by_book(book_id)
        
        with open(output_path, 'w', encoding=config.EXPORT_ENCODING, newline='') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write headers
            writer.writerow(self.headers)
            
            # Write data
            for chapter in chapters:
                content_list = db.get_generated_content_by_chapter(chapter['id'])
                
                for content in content_list:
                    import json
                    
                    # Format options
                    options_str = ''
                    if content.get('options_json'):
                        try:
                            options = json.loads(content['options_json'])
                            options_str = '\n'.join(options)
                        except:
                            options_str = content['options_json']
                    
                    # Format status
                    status_cn = {
                        'pending': '待生成',
                        'generated': '已生成',
                        'verified': '已校验'
                    }.get(content['status'], content['status'])
                    
                    # Content type
                    content_type_cn = '问答' if content['content_type'] == 'qa' else '习题'
                    
                    # Content excerpt
                    content_excerpt = chapter.get('content_md', '')[:100] + '...' if chapter.get('content_md') else ''
                    
                    # Model info
                    model_info = f"{content['model_name']}"
                    if content.get('model_version'):
                        model_info += f" ({content['model_version']})"
                    
                    row = [
                        chapter['id'],
                        chapter['title'],
                        content_excerpt,
                        content_type_cn,
                        content['question'],
                        options_str,
                        content['answer'],
                        content.get('explanation', ''),
                        model_info,
                        status_cn
                    ]
                    
                    writer.writerow(row)
        
        return output_path


# Global exporter instance
excel_exporter = ExcelExporter()
